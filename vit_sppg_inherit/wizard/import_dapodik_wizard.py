
import base64
import csv
import io
import logging
from datetime import datetime
from collections import defaultdict
from odoo import models, fields, api, _
from odoo.exceptions import ValidationError
from odoo.tools.translate import _

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

_logger = logging.getLogger(__name__)

class ImportDapodikWizard(models.TransientModel):
    _name = 'vit.import.dapodik.wizard'
    _description = 'Wizard to Import Dapodik Data'

    file_data = fields.Binary(string='File', required=True)
    file_name = fields.Char(string='File Name')
    # Removed fields to avoid database upgrade issues
    # batch_size and update_mode will be hardcoded for now

    def _to_int(self, value):
        """Helper to safely convert string to integer."""
        try:
            return int(float(value)) if value else 0
        except (ValueError, TypeError):
            return 0

    def _to_float(self, value):
        """Helper to safely convert string to float."""
        try:
            return float(value) if value else 0.0
        except (ValueError, TypeError):
            return 0.0

    def _detect_file_type(self, raw_bytes):
        """Detect file type based on file signature"""
        if raw_bytes.startswith(b'PK'):
            return 'xlsx'
        elif raw_bytes.startswith(b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'):
            return 'xls'
        else:
            return 'csv'

    def _to_float(self, value):
        """Helper to safely convert string to float."""
        try:
            return float(value) if value else 0.0
        except (ValueError, TypeError):
            return 0.0

    def _normalize_name(self, name):
        """Normalize name for consistent matching"""
        if not name:
            return ""
        
        # Convert to uppercase
        normalized = str(name).strip().upper()
        
        # Remove "PROV.", "PROVINSI", "PROP.", etc.
        prefixes_to_remove = ['PROV.', 'PROVINSI', 'PROP.', 'PROV', 'PROP']
        for prefix in prefixes_to_remove:
            if normalized.startswith(prefix):
                normalized = normalized[len(prefix):].strip()
                break
        
        # Remove dots and normalize spacing
        normalized = normalized.replace('.', '')
        
        # Handle specific cases
        name_mappings = {
            'DKI JAKARTA': 'DKI JAKARTA',
            'D K I JAKARTA': 'DKI JAKARTA', 
            'DI YOGYAKARTA': 'DI YOGYAKARTA',
            'D I YOGYAKARTA': 'DI YOGYAKARTA',
            'KEP BANGKA BELITUNG': 'KEPULAUAN BANGKA BELITUNG',
            'BANGKA BELITUNG': 'KEPULAUAN BANGKA BELITUNG',
            'KEP RIAU': 'KEPULAUAN RIAU',
            'SULAWESI UTARA': 'SULAWESI UTARA',
            'SULUT': 'SULAWESI UTARA',
            'SULAWESI SELATAN': 'SULAWESI SELATAN', 
            'SULSEL': 'SULAWESI SELATAN',
            'SULAWESI TENGAH': 'SULAWESI TENGAH',
            'SULTENG': 'SULAWESI TENGAH',
            'SULAWESI TENGGARA': 'SULAWESI TENGGARA',
            'SULTRA': 'SULAWESI TENGGARA',
            'SULAWESI BARAT': 'SULAWESI BARAT',
            'SULBAR': 'SULAWESI BARAT',
            'KALIMANTAN UTARA': 'KALIMANTAN UTARA',
            'KALUT': 'KALIMANTAN UTARA',
            'KALIMANTAN TIMUR': 'KALIMANTAN TIMUR',
            'KALTIM': 'KALIMANTAN TIMUR',
            'KALIMANTAN SELATAN': 'KALIMANTAN SELATAN',
            'KALSEL': 'KALIMANTAN SELATAN',
            'KALIMANTAN TENGAH': 'KALIMANTAN TENGAH',
            'KALTENG': 'KALIMANTAN TENGAH',
            'KALIMANTAN BARAT': 'KALIMANTAN BARAT',
            'KALBAR': 'KALIMANTAN BARAT',
            'NUSA TENGGARA BARAT': 'NUSA TENGGARA BARAT',
            'NTB': 'NUSA TENGGARA BARAT',
            'NUSA TENGGARA TIMUR': 'NUSA TENGGARA TIMUR', 
            'NTT': 'NUSA TENGGARA TIMUR',
            'PAPUA BARAT': 'PAPUA BARAT',
            'PAPBAR': 'PAPUA BARAT',
            'PAPUA BARAT DAYA': 'PAPUA BARAT DAYA',
            'PAPUA TENGAH': 'PAPUA TENGAH',
            'PAPUA PEGUNUNGAN': 'PAPUA PEGUNUNGAN',
            'PAPUA SELATAN': 'PAPUA SELATAN'
        }
        
        # Check if normalized name exists in mapping
        if normalized in name_mappings:
            return name_mappings[normalized]
            
        # Clean up multiple spaces
        normalized = ' '.join(normalized.split())
        
        return normalized# -*- coding: utf-8 -*-

    def _read_excel_file(self, raw_bytes, file_type):
        """Read Excel file and convert to list of rows - OPTIMIZED"""
        if file_type == 'xlsx':
            if not HAS_OPENPYXL:
                raise ValidationError(_("openpyxl library is required"))
            
            # Use read_only mode for better memory efficiency
            workbook = openpyxl.load_workbook(
                io.BytesIO(raw_bytes), 
                read_only=True,  # Memory optimization
                data_only=True   # Get values instead of formulas
            )
            worksheet = workbook.active
            
            # Generator approach to save memory
            for row in worksheet.iter_rows(values_only=True):
                yield [str(cell) if cell is not None else '' for cell in row]
            
            workbook.close()
            
        elif file_type == 'xls':
            if not HAS_XLRD:
                raise ValidationError(_("xlrd library is required"))
            
            workbook = xlrd.open_workbook(file_contents=raw_bytes, on_demand=True)
            worksheet = workbook.sheet_by_index(0)
            
            for row_idx in range(worksheet.nrows):
                row = []
                for col_idx in range(worksheet.ncols):
                    cell = worksheet.cell(row_idx, col_idx)
                    if cell.ctype == xlrd.XL_CELL_NUMBER:
                        if cell.value == int(cell.value):
                            row.append(str(int(cell.value)))
                        else:
                            row.append(str(cell.value))
                    elif cell.ctype == xlrd.XL_CELL_DATE:
                        row.append(str(cell.value))
                    else:
                        row.append(str(cell.value))
                yield row

    def _read_csv_file(self, raw_bytes):
        """Read CSV file and convert to generator - OPTIMIZED"""
        cleaned_bytes = raw_bytes.replace(b'\0', b'')
        
        text_stream = None
        for encoding in ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']:
            try:
                text_stream = io.TextIOWrapper(
                    io.BytesIO(cleaned_bytes), 
                    encoding=encoding,
                    errors='replace'
                )
                pos = text_stream.tell()
                test_line = text_stream.readline()
                text_stream.seek(pos)
                
                if test_line:
                    break
            except Exception:
                if text_stream:
                    text_stream.close()
                    text_stream = None
                continue
        
        if not text_stream:
            raise ValidationError(_("Could not decode CSV file"))
        
        # Auto-detect delimiter
        text_stream.seek(0)
        sample = text_stream.read(1024)
        text_stream.seek(0)
        
        delimiter = ','
        for test_delimiter in [',', ';', '\t', '|']:
            if sample.count(test_delimiter) > sample.count(delimiter):
                delimiter = test_delimiter
        
        # Use generator for memory efficiency
        reader = csv.reader(text_stream, delimiter=delimiter, quotechar='"')
        for row in reader:
            yield row
        
        text_stream.close()

    def _bulk_load_existing_data(self):
        """Pre-load existing data to minimize database queries"""
        _logger.info("Pre-loading existing data...")
        
        # Load all existing data in memory for fast lookups
        existing_data = {
            'provinsi': {},  # name -> record
            'kab_kota': {},  # (name, provinsi_id) -> record  
            'dapodik': {}    # npsn -> record
        }
        
        # Load provinsi
        provinsi_records = self.env['vit.master_provinsi'].search([])
        for rec in provinsi_records:
            normalized_name = self._normalize_name(rec.name)
            existing_data['provinsi'][normalized_name] = rec
        
        # Load kab/kota
        kab_kota_records = self.env['vit.master_kab_kota'].search([])
        for rec in kab_kota_records:
            normalized_name = self._normalize_name(rec.name)
            key = (normalized_name, rec.provinsi_id.id)
            existing_data['kab_kota'][key] = rec
        
        # Load dapodik
        dapodik_records = self.env['vit.master_dapodik'].search([])
        for rec in dapodik_records:
            existing_data['dapodik'][rec.npsn] = rec
            
        _logger.info(f"Loaded {len(provinsi_records)} provinsi, {len(kab_kota_records)} kab/kota, {len(dapodik_records)} dapodik records")
        return existing_data

    def _process_batch(self, batch_data, existing_data):
        """Process a batch of records efficiently"""
        # Hardcoded settings to avoid database upgrade issues
        batch_size = 1000  # Can be adjusted as needed
        update_mode = 'create_update'  # create_only, update_only, or create_update
        
        provinsi_to_create = []
        kab_kota_to_create = []
        dapodik_to_create = []
        dapodik_to_update = []
        
        batch_stats = {
            'processed': 0,
            'created': 0, 
            'updated': 0,
            'skipped': 0,
            'errors': 0
        }

        for row_data in batch_data:
            try:
                row_number, row = row_data
                
                # Validate and extract data
                if not row or len(row) < 19:
                    batch_stats['skipped'] += 1
                    continue
                
                npsn = str(row[4]).strip()
                name = str(row[5]).strip() 
                provinsi_name = self._normalize_name(str(row[1]).strip())
                kab_kota_name = self._normalize_name(str(row[2]).strip())
                
                if not all([npsn, name, provinsi_name, kab_kota_name]):
                    batch_stats['skipped'] += 1
                    continue

                # --- Process Provinsi ---
                provinsi_key = provinsi_name  # Already normalized
                provinsi = existing_data['provinsi'].get(provinsi_key)
                
                if not provinsi:
                    # Check if already queued for creation in this batch
                    existing_queued = next((p for p in provinsi_to_create if self._normalize_name(p['name']) == provinsi_key), None)
                    if not existing_queued:
                        # Store original name for database, but use normalized for lookup
                        original_name = str(row[1]).strip()  # Keep original formatting for database
                        provinsi_data = {'name': original_name}
                        provinsi_to_create.append(provinsi_data)
                        # Create temporary record for this batch
                        provinsi = type('obj', (object,), {'id': f'temp_{len(provinsi_to_create)}', 'name': original_name})()
                        existing_data['provinsi'][provinsi_key] = provinsi
                    else:
                        provinsi = existing_data['provinsi'][provinsi_key]

                # --- Process Kab/Kota ---
                kab_kota_key = (kab_kota_name, provinsi.id)  # Already normalized
                kab_kota = existing_data['kab_kota'].get(kab_kota_key)
                
                if not kab_kota:
                    existing_queued = next((k for k in kab_kota_to_create 
                                         if self._normalize_name(k['name']) == kab_kota_name 
                                         and k['provinsi_id'] == provinsi.id), None)
                    if not existing_queued:
                        # Store original name for database
                        original_kab_kota_name = str(row[2]).strip()
                        kab_kota_data = {
                            'name': original_kab_kota_name,
                            'provinsi_id': provinsi.id
                        }
                        kab_kota_to_create.append(kab_kota_data)
                        # Create temporary record for this batch
                        kab_kota = type('obj', (object,), {'id': f'temp_{len(kab_kota_to_create)}', 'name': original_kab_kota_name})()
                        existing_data['kab_kota'][kab_kota_key] = kab_kota
                    else:
                        kab_kota = existing_data['kab_kota'][kab_kota_key]

                # --- Process Dapodik ---
                dapodik_vals = {
                    'name': name,
                    'npsn': npsn,
                    'status_kepemilikan': str(row[6]).strip(),
                    'bentuk_pendidikan': str(row[7]).strip(),
                    'status_sekolah': str(row[8]).strip(),
                    'alamat': str(row[9]).strip(),
                    'lintang': self._to_float(row[10]),
                    'bujur': self._to_float(row[11]),
                    'jumlah_pd': self._to_int(row[12]),
                    'jumlah_guru': self._to_int(row[13]),
                    'jumlah_tendik': self._to_int(row[14]),
                    'update_jumlah_pd': self._to_int(row[16]),
                    'update_jumlah_tendik': self._to_int(row[18]),
                    'provinsi_id': provinsi.id,
                   'kab_kota_id': kab_kota.id,
                }
                existing_dapodik = existing_data['dapodik'].get(npsn)
                
                if existing_dapodik and update_mode in ['update_only', 'create_update']:
                    dapodik_vals['id'] = existing_dapodik.id
                    dapodik_to_update.append(dapodik_vals)
                    batch_stats['updated'] += 1
                elif not existing_dapodik and update_mode in ['create_only', 'create_update']:
                    dapodik_to_create.append(dapodik_vals)
                    batch_stats['created'] += 1
                else:
                    batch_stats['skipped'] += 1
                    
                batch_stats['processed'] += 1
                
            except Exception as e:
                _logger.error(f"Row {row_number}: Error processing: {e}")
                batch_stats['errors'] += 1

        # Execute bulk operations
        try:
            # Create missing provinsi
            if provinsi_to_create:
                created_provinsi = self.env['vit.master_provinsi'].create(provinsi_to_create)
                _logger.info(f"Bulk created {len(created_provinsi)} provinsi")
                
                # Update cache with real IDs
                for i, provinsi in enumerate(created_provinsi):
                    normalized_key = self._normalize_name(provinsi.name)
                    existing_data['provinsi'][normalized_key] = provinsi
            
            # Create missing kab/kota
            if kab_kota_to_create:
                # Update provinsi_id for kab_kota if needed
                for kab_kota_data in kab_kota_to_create:
                    if isinstance(kab_kota_data['provinsi_id'], str):
                        # Find real provinsi_id
                        provinsi_name = next(p['name'] for p in provinsi_to_create 
                                           if f"temp_{provinsi_to_create.index(p)+1}" == kab_kota_data['provinsi_id'])
                        real_provinsi = existing_data['provinsi'][provinsi_name.lower()]
                        kab_kota_data['provinsi_id'] = real_provinsi.id
                        
                created_kab_kota = self.env['vit.master_kab_kota'].create(kab_kota_to_create)
                _logger.info(f"Bulk created {len(created_kab_kota)} kab/kota")
                
                # Update cache
                for kab_kota in created_kab_kota:
                    normalized_key = self._normalize_name(kab_kota.name)
                    key = (normalized_key, kab_kota.provinsi_id.id)
                    existing_data['kab_kota'][key] = kab_kota
            
            # Update dapodik provinsi_id and kab_kota_id if needed
            for dapodik_data in dapodik_to_create + dapodik_to_update:
                if isinstance(dapodik_data['provinsi_id'], str):
                    # Find real provinsi
                    for provinsi in existing_data['provinsi'].values():
                        if hasattr(provinsi, 'id') and not isinstance(provinsi.id, str):
                            dapodik_data['provinsi_id'] = provinsi.id
                            break
                            
                if isinstance(dapodik_data['kab_kota_id'], str):
                    # Find real kab_kota
                    for kab_kota in existing_data['kab_kota'].values():
                        if hasattr(kab_kota, 'id') and not isinstance(kab_kota.id, str):
                            dapodik_data['kab_kota_id'] = kab_kota.id
                            break
            
            # Bulk create dapodik
            if dapodik_to_create:
                created_dapodik = self.env['vit.master_dapodik'].create(dapodik_to_create)
                _logger.info(f"Bulk created {len(created_dapodik)} dapodik")
                
                # Update cache
                for dapodik in created_dapodik:
                    existing_data['dapodik'][dapodik.npsn] = dapodik
            
            # Bulk update dapodik 
            if dapodik_to_update:
                for dapodik_data in dapodik_to_update:
                    dapodik_id = dapodik_data.pop('id')
                    self.env['vit.master_dapodik'].browse(dapodik_id).write(dapodik_data)
                _logger.info(f"Bulk updated {len(dapodik_to_update)} dapodik")
                
        except Exception as e:
            _logger.error(f"Bulk operation failed: {e}")
            batch_stats['errors'] += batch_stats['processed']
            batch_stats['processed'] = 0
            batch_stats['created'] = 0
            batch_stats['updated'] = 0

        return batch_stats

    def action_import(self):
        if not self.file_data:
            raise ValidationError(_("Please upload a file to import."))

        # Initialize counters
        total_stats = {
            'processed': 0,
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'errors': 0,
            'total_rows': 0
        }

        try:
            # 1. Decode and detect file
            raw_bytes = base64.b64decode(self.file_data)
            file_type = self._detect_file_type(raw_bytes)
            _logger.info(f"Processing {file_type.upper()} file ({len(raw_bytes)} bytes)")

            # 2. Pre-load existing data
            existing_data = self._bulk_load_existing_data()

            # 3. Read file as generator
            if file_type in ['xlsx', 'xls']:
                rows_generator = self._read_excel_file(raw_bytes, file_type)
            else:
                rows_generator = self._read_csv_file(raw_bytes)

            # 4. Process in batches
            batch_data = []
            row_number = 0
            header_skipped = False
            batch_size = 1000  # Hardcoded batch size
            
            for row in rows_generator:
                row_number += 1
                
                # Skip header
                if not header_skipped:
                    header_skipped = True
                    continue
                
                total_stats['total_rows'] += 1
                batch_data.append((row_number, row))
                
                # Process batch when full
                if len(batch_data) >= batch_size:
                    _logger.info(f"Processing batch {total_stats['total_rows']//batch_size} ({len(batch_data)} records)")
                    
                    batch_stats = self._process_batch(batch_data, existing_data)
                    
                    # Update totals
                    for key in batch_stats:
                        total_stats[key] += batch_stats[key]
                    
                    # Commit batch
                    self.env.cr.commit()
                    
                    # Clear batch
                    batch_data = []
                    
                    _logger.info(f"Batch completed. Total processed: {total_stats['processed']}")

            # Process remaining records
            if batch_data:
                _logger.info(f"Processing final batch ({len(batch_data)} records)")
                batch_stats = self._process_batch(batch_data, existing_data)
                
                for key in batch_stats:
                    total_stats[key] += batch_stats[key]
                
                self.env.cr.commit()

        except Exception as e:
            error_message = f"Import failed: {str(e)}"
            _logger.error(error_message)
            raise ValidationError(error_message)

        # Success message
        success_message = (
            f"Import completed!\n"
            f"File: {file_type.upper()} ({len(raw_bytes)//1024//1024}MB)\n"
            f"Mode: Create and Update\n"
            f"Total rows: {total_stats['total_rows']:,}\n"
            f"Processed: {total_stats['processed']:,}\n"
            f"Created: {total_stats['created']:,}\n"
            f"Updated: {total_stats['updated']:,}\n" 
            f"Skipped: {total_stats['skipped']:,}\n"
            f"Errors: {total_stats['errors']:,}"
        )
        
        _logger.info(success_message)

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Import Completed'),
                'message': success_message,
                'type': 'success' if total_stats['processed'] > 0 else 'warning',
                'sticky': True,
            }
        }